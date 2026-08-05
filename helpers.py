import pandas as pd
from datetime import datetime
import io
from openpyxl.styles import numbers
from models import BalanceReport, BalanceReportFL
from deposit_report import DepositReportGenerator

# ============================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ДЛЯ ФОРМАТИРОВАНИЯ
# ============================================

def format_number(value: float, with_sign: bool = False) -> str:
    """
    Форматирует число в формат: 50584212,94
    Если with_sign=True: +50584212,94 или -50584212,94
    """
    if pd.isna(value) or value == 0:
        return ""
    if with_sign:
        return f"{value:+.2f}".replace(".", ",")
    return f"{value:.2f}".replace(".", ",")

def excel_date(date_val) -> int:
    """
    Преобразует дату в числовой формат Excel.
    Поддерживает: строки "31.07.2026", datetime-объекты, Timestamp.
    """
    if pd.isna(date_val) or date_val == "":
        return None
    
    if isinstance(date_val, (pd.Timestamp, datetime)):
        return (date_val - pd.Timestamp("1899-12-30")).days
    
    if isinstance(date_val, str):
        date_val = date_val.strip()
        try:
            dt = pd.to_datetime(date_val, format="%d.%m.%Y", errors="coerce")
            if pd.isna(dt):
                dt = pd.to_datetime(date_val, errors="coerce")
            if pd.isna(dt):
                return None
            return (dt - pd.Timestamp("1899-12-30")).days
        except:
            return None
    
    return None

def apply_number_format(worksheet, col_letter, start_row=2, format_str='# ###0.00'):
    """Применяет числовой формат к колонке"""
    for row in range(start_row, worksheet.max_row + 1):
        cell = worksheet[f"{col_letter}{row}"]
        cell.number_format = format_str

def apply_date_format(worksheet, col_letter, start_row=2):
    """Применяет формат даты к колонке"""
    for row in range(start_row, worksheet.max_row + 1):
        cell = worksheet[f"{col_letter}{row}"]
        cell.number_format = 'DD.MM.YYYY'

# ============================================
# ЭКСПОРТ В EXCEL
# ============================================

def create_excel_report(
    ip_report: BalanceReport,
    ip_operations: pd.DataFrame,
    fl_report: BalanceReportFL = None,
    fl_operations: pd.DataFrame = None
) -> io.BytesIO:
    """Создает единый Excel-файл с отчетами"""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # ============================================
        # ИП_Динамика
        # ============================================
        if ip_report is not None and not ip_report.monthly_dynamics.empty:
            df_dynamics = ip_report.monthly_dynamics.copy()
            
            balances = []
            for i in range(len(df_dynamics)):
                if i == 0:
                    start_bal = ip_report.start_balance
                else:
                    start_bal = df_dynamics.iloc[i-1]["balance"]
                balances.append(start_bal)
            
            df_dynamics["Баланс начало месяца"] = balances
            df_dynamics["Баланс конец месяца"] = df_dynamics["balance"]
            df_dynamics = df_dynamics.rename(columns={"month": "Месяц"})
            
            result_df = df_dynamics[["Месяц", "Баланс начало месяца", "Баланс конец месяца"]].copy()
            # Сохраняем как числа (без форматирования текстом)
            result_df["Баланс начало месяца"] = result_df["Баланс начало месяца"]
            result_df["Баланс конец месяца"] = result_df["Баланс конец месяца"]
            
            result_df.to_excel(writer, sheet_name="ИП_Динамика", index=False)
            
            # Применяем числовой формат к колонкам B и C
            worksheet = writer.sheets["ИП_Динамика"]
            for col in ['B', 'C']:
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet[f"{col}{row}"]
                    cell.number_format = '# ###0.00'
        
        # ============================================
        # ИП_Операции
        # ============================================
        if ip_operations is not None and not ip_operations.empty:
            ops_df = ip_operations.copy()
            
            ops_df["Дата"] = ops_df["date"].apply(excel_date)
            ops_df["Дебет"] = ops_df["debit"]
            ops_df["Кредит"] = ops_df["credit"]
            ops_df["Итого"] = ops_df["amount"]
            ops_df["Описание"] = ops_df["description"]
            
            result_ops = ops_df[["Дата", "Дебет", "Кредит", "Итого", "Описание"]].copy()
            result_ops.to_excel(writer, sheet_name="ИП_Операции", index=False)
            
            worksheet = writer.sheets["ИП_Операции"]
            # Дата — формат даты
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet[f"A{row}"]
                cell.number_format = 'DD.MM.YYYY'
            # Числа — числовой формат
            for col in ['B', 'C', 'D']:
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet[f"{col}{row}"]
                    cell.number_format = '# ###0.00'
        
        # ============================================
        # Депозиты_Динамика
        # ============================================
        if ip_operations is not None and not ip_operations.empty:
            deposit_ops = ip_operations.attrs.get("deposits", pd.DataFrame())
            
            if not deposit_ops.empty:
                deposit_report = DepositReportGenerator.generate_report(deposit_ops)
                
                if not deposit_report.empty:
                    deposit_report_copy = deposit_report.copy()
                    
                    if "Дата начала" in deposit_report_copy.columns:
                        deposit_report_copy["Дата начала"] = deposit_report_copy["Дата начала"].apply(
                            lambda x: excel_date(x) if pd.notna(x) else ""
                        )
                    
                    if "Дата завершения" in deposit_report_copy.columns:
                        deposit_report_copy["Дата завершения"] = deposit_report_copy["Дата завершения"].apply(
                            lambda x: excel_date(x) if pd.notna(x) else ""
                        )
                    
                    deposit_report_copy["Сумма депозита (руб)"] = deposit_report_copy["Сумма депозита (руб)"]
                    deposit_report_copy["Процент депозита (руб)"] = deposit_report_copy["Процент депозита (руб)"]
                    
                    deposit_report_copy.to_excel(writer, sheet_name="Депозиты_Динамика", index=False)
                    
                    worksheet = writer.sheets["Депозиты_Динамика"]
                    # Даты
                    for col in ['B', 'C']:
                        for row in range(2, worksheet.max_row + 1):
                            cell = worksheet[f"{col}{row}"]
                            if cell.value:
                                cell.number_format = 'DD.MM.YYYY'
                    # Числа
                    for col in ['D', 'E']:
                        for row in range(2, worksheet.max_row + 1):
                            cell = worksheet[f"{col}{row}"]
                            cell.number_format = '# ###0.00'
        
        # ============================================
        # Депозиты_Операции
        # ============================================
        if ip_operations is not None and not ip_operations.empty:
            deposit_ops = ip_operations.attrs.get("deposits", pd.DataFrame())
            
            if not deposit_ops.empty:
                detail_df = deposit_ops.copy()
                
                from deposit_report import DepositReportGenerator as DRG
                detail_df["Номер сделки"] = detail_df["description"].apply(DRG.extract_deal_number)
                detail_df["Дата"] = detail_df["date"].apply(excel_date)
                detail_df["Сумма"] = detail_df["amount"]
                detail_df["Назначение платежа"] = detail_df["description"]
                
                result_detail = detail_df[["Номер сделки", "Дата", "Сумма", "Назначение платежа"]].copy()
                result_detail = result_detail.dropna(subset=["Номер сделки"])
                
                result_detail.to_excel(writer, sheet_name="Депозиты_Операции", index=False)
                
                worksheet = writer.sheets["Депозиты_Операции"]
                # Дата
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet[f"B{row}"]
                    if cell.value:
                        cell.number_format = 'DD.MM.YYYY'
                # Сумма
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet[f"C{row}"]
                    cell.number_format = '# ###0.00'
        
        # ============================================
        # ФЛ_Динамика
        # ============================================
        if fl_report is not None and not fl_report.monthly_dynamics.empty:
            df_fl_dynamics = fl_report.monthly_dynamics.copy()
            
            balances = []
            for i in range(len(df_fl_dynamics)):
                if i == 0:
                    start_bal = fl_report.start_balance
                else:
                    start_bal = df_fl_dynamics.iloc[i-1]["balance"]
                balances.append(start_bal)
            
            df_fl_dynamics["Баланс начало месяца"] = balances
            df_fl_dynamics["Баланс конец месяца"] = df_fl_dynamics["balance"]
            df_fl_dynamics = df_fl_dynamics.rename(columns={"month": "Месяц"})
            
            result_df = df_fl_dynamics[["Месяц", "Баланс начало месяца", "Баланс конец месяца"]].copy()
            result_df.to_excel(writer, sheet_name="ФЛ_Динамика", index=False)
            
            worksheet = writer.sheets["ФЛ_Динамика"]
            for col in ['B', 'C']:
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet[f"{col}{row}"]
                    cell.number_format = '# ###0.00'
        else:
            empty_df = pd.DataFrame({"Сообщение": ["Нет данных для динамики ФЛ"]})
            empty_df.to_excel(writer, sheet_name="ФЛ_Динамика", index=False)
        
        # ============================================
        # ФЛ_Операции
        # ============================================
        if fl_operations is not None and not fl_operations.empty:
            fl_ops = fl_operations.copy()
            
            fl_ops["Дата"] = fl_ops["date"].apply(excel_date)
            fl_ops["Описание"] = fl_ops["description"]
            fl_ops["Сумма"] = fl_ops["amount"]
            fl_ops["Счет"] = fl_ops["account_name"] if "account_name" in fl_ops.columns else ""
            fl_ops["Тип"] = fl_ops["type"] if "type" in fl_ops.columns else ""
            fl_ops["Категория"] = fl_ops["category"] if "category" in fl_ops.columns else ""
            
            result_fl = fl_ops[["Дата", "Описание", "Сумма", "Счет", "Тип", "Категория"]].copy()
            result_fl.to_excel(writer, sheet_name="ФЛ_Операции", index=False)
            
            worksheet = writer.sheets["ФЛ_Операции"]
            # Дата
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet[f"A{row}"]
                if cell.value:
                    cell.number_format = 'DD.MM.YYYY'
            # Сумма
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet[f"C{row}"]
                cell.number_format = '# ###0.00'
        else:
            empty_df = pd.DataFrame({"Сообщение": ["Нет данных по операциям ФЛ"]})
            empty_df.to_excel(writer, sheet_name="ФЛ_Операции", index=False)
        
        # ============================================
        # ФЛ_Вклады
        # ============================================
        if fl_report is not None and fl_report.deposits_data:
            all_deposits = []
            for deposit_item in fl_report.deposits_data:
                df_dep = deposit_item["data"].copy()
                df_dep["Название счета"] = deposit_item["account_name"]
                all_deposits.append(df_dep)
            
            if all_deposits:
                combined_df = pd.concat(all_deposits, ignore_index=True)
                combined_df = combined_df[["Название счета", "month", "balance", "interest"]]
                combined_df = combined_df.rename(columns={
                    "month": "Месяц",
                    "balance": "Остаток на конец месяца",
                    "interest": "Выплата процентов"
                })
                combined_df.to_excel(writer, sheet_name="ФЛ_Вклады", index=False)
                
                worksheet = writer.sheets["ФЛ_Вклады"]
                for col in ['C', 'D']:
                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet[f"{col}{row}"]
                        cell.number_format = '# ###0.00'
            else:
                empty_df = pd.DataFrame({"Сообщение": ["Нет данных по вкладам ФЛ"]})
                empty_df.to_excel(writer, sheet_name="ФЛ_Вклады", index=False)
        else:
            empty_df = pd.DataFrame({"Сообщение": ["Нет данных по вкладам ФЛ"]})
            empty_df.to_excel(writer, sheet_name="ФЛ_Вклады", index=False)
        
        # ============================================
        # ФЛ_Вклады_Операции
        # ============================================
        if fl_operations is not None and not fl_operations.empty:
            deposit_names = ["Альфа-Счёт на минимальный остаток", "Альфа-Счёт на ежедневный остаток"]
            fl_deposit_ops = fl_operations[fl_operations["account_name"].isin(deposit_names)].copy()
            
            if not fl_deposit_ops.empty:
                fl_deposit_ops["Дата"] = fl_deposit_ops["date"].apply(excel_date)
                fl_deposit_ops["Описание"] = fl_deposit_ops["description"]
                fl_deposit_ops["Сумма"] = fl_deposit_ops["amount"]
                fl_deposit_ops["Счет"] = fl_deposit_ops["account_name"]
                fl_deposit_ops["Тип"] = fl_deposit_ops["type"] if "type" in fl_deposit_ops.columns else ""
                
                result_deposit_ops = fl_deposit_ops[["Дата", "Описание", "Сумма", "Счет", "Тип"]].copy()
                result_deposit_ops.to_excel(writer, sheet_name="ФЛ_Вклады_Операции", index=False)
                
                worksheet = writer.sheets["ФЛ_Вклады_Операции"]
                # Дата
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet[f"A{row}"]
                    if cell.value:
                        cell.number_format = 'DD.MM.YYYY'
                # Сумма
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet[f"C{row}"]
                    cell.number_format = '# ###0.00'
            else:
                empty_df = pd.DataFrame({"Сообщение": ["Нет операций по вкладам ФЛ"]})
                empty_df.to_excel(writer, sheet_name="ФЛ_Вклады_Операции", index=False)
        else:
            empty_df = pd.DataFrame({"Сообщение": ["Нет данных по вкладам ФЛ"]})
            empty_df.to_excel(writer, sheet_name="ФЛ_Вклады_Операции", index=False)
    
    output.seek(0)
    return output


def export_deposit_report_to_excel(report_df: pd.DataFrame, ip_operations: pd.DataFrame) -> io.BytesIO:
    """
    Устаревшая функция. Оставлена для обратной совместимости.
    """
    deposit_ops = ip_operations.attrs.get("deposits", pd.DataFrame())
    return DepositReportGenerator.export_to_excel(report_df, deposit_ops)
